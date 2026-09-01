from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from negocio.models import CashEntry, Customer, Ingredient, Product, Receivable, Recipe, RecipeItem


def decimal_value(value, default=Decimal("0")):
    if value in (None, "", "-"):
        return default
    try:
        return Decimal(str(value).replace("R$", "").replace(" ", "").replace(",", "."))
    except (InvalidOperation, ValueError):
        return default


def date_value(value):
    if isinstance(value, datetime):
        return value.date()
    if not value:
        return None
    for pattern in ("%d/%m/%Y", "%d/%m%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(value), pattern).date()
        except ValueError:
            continue
    return None


class Command(BaseCommand):
    help = "Importa custos, receitas, clientes, contas a receber e caixa da planilha enviada."

    def add_arguments(self, parser):
        parser.add_argument("arquivo", help="Caminho do arquivo .xlsx")

    @transaction.atomic
    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError("Instale openpyxl com: pip install -r requirements.txt") from exc

        path = options["arquivo"]
        try:
            formulas = load_workbook(path, data_only=False)
            values = load_workbook(path, data_only=True)
        except OSError as exc:
            raise CommandError(f"Não foi possível abrir a planilha: {exc}") from exc

        costs = values["Tabela de Custos"]
        ingredients_by_name = {}
        for row in range(10, 308):
            name = costs.cell(row, 1).value
            if not name:
                continue
            name = str(name).strip()
            quantity = decimal_value(costs.cell(row, 2).value)
            price = decimal_value(costs.cell(row, 3).value)
            if quantity <= 0:
                continue
            unit = Ingredient.Unit.UNIT if name.lower() in {"ovos", "marshmellow grande", "marshmallow mini"} else Ingredient.Unit.GRAM
            ingredient, _ = Ingredient.objects.update_or_create(
                name=name,
                defaults={"unit": unit, "package_quantity": quantity, "package_price": price},
            )
            ingredients_by_name[name.casefold()] = ingredient

        recipes_sheet = formulas["Receitas de doces"]
        recipes_values = values["Receitas de doces"]
        for header_row in range(3, 498, 26):
            for ingredient_col in (1, 5, 9, 13, 17):
                name = recipes_sheet.cell(header_row, ingredient_col + 1).value
                if not name:
                    continue
                name = str(name).strip()
                value_col = ingredient_col + 2
                yield_quantity = int(decimal_value(recipes_values.cell(header_row + 18, value_col).value, Decimal("1")) or 1)
                sale_price = decimal_value(recipes_values.cell(header_row + 21, value_col).value)
                if sale_price <= 0:
                    sale_price = decimal_value(recipes_values.cell(header_row + 20, value_col).value, Decimal("0.01"))
                recipe, _ = Recipe.objects.update_or_create(
                    name=name,
                    defaults={
                        "yield_quantity": max(1, yield_quantity),
                        "extra_cost": decimal_value(recipes_values.cell(header_row + 14, value_col).value),
                    },
                )
                recipe.items.all().delete()
                for row in range(header_row + 2, header_row + 15):
                    ingredient_name = recipes_sheet.cell(row, ingredient_col).value
                    quantity = decimal_value(recipes_sheet.cell(row, ingredient_col + 1).value)
                    if not ingredient_name or quantity <= 0:
                        continue
                    key = str(ingredient_name).strip().casefold()
                    ingredient = ingredients_by_name.get(key)
                    if not ingredient:
                        ingredient = Ingredient.objects.filter(name__iexact=str(ingredient_name).strip()).first()
                    if not ingredient:
                        ingredient = Ingredient.objects.create(name=str(ingredient_name).strip())
                        ingredients_by_name[key] = ingredient
                    RecipeItem.objects.update_or_create(recipe=recipe, ingredient=ingredient, defaults={"quantity": quantity})
                Product.objects.update_or_create(
                    name=name,
                    defaults={"recipe": recipe, "sale_price": max(sale_price, Decimal("0.01")), "active": True},
                )

        clients = values["Clientes"]
        for row in range(10, clients.max_row + 1):
            name = clients.cell(row, 2).value
            if not name:
                continue
            phone = clients.cell(row, 5).value or clients.cell(row, 4).value or f"sem-telefone-{row}"
            Customer.objects.update_or_create(
                phone=str(phone).strip(),
                defaults={
                    "name": str(name).strip(),
                    "address": str(clients.cell(row, 3).value or ""),
                    "email": str(clients.cell(row, 6).value or ""),
                    "birthday": date_value(clients.cell(row, 7).value),
                },
            )

        receivables = values["Contas a Receber"]
        for row in range(7, receivables.max_row + 1):
            name = receivables.cell(row, 1).value
            amount = decimal_value(receivables.cell(row, 4).value)
            if not name or amount <= 0:
                continue
            customer = Customer.objects.filter(name__iexact=str(name).strip()).first()
            if not customer:
                customer = Customer.objects.create(name=str(name).strip(), phone=f"importado-{row}")
            status = str(receivables.cell(row, 5).value or "PENDENTE").upper()
            Receivable.objects.get_or_create(
                customer=customer,
                description=str(receivables.cell(row, 2).value or "Venda importada"),
                amount=amount,
                defaults={
                    "due_date": date_value(receivables.cell(row, 3).value),
                    "status": Receivable.Status.PAID if status == "PAGO" else Receivable.Status.PENDING,
                    "paid_at": date_value(receivables.cell(row, 6).value),
                },
            )

        cash = values["Fluxo de Caixa"]
        for row in range(9, cash.max_row + 1):
            date = date_value(cash.cell(row, 1).value)
            description = cash.cell(row, 2).value
            income = decimal_value(cash.cell(row, 3).value)
            expense = decimal_value(cash.cell(row, 4).value)
            if not description or (income <= 0 and expense <= 0):
                continue
            kind = CashEntry.Kind.INCOME if income > 0 else CashEntry.Kind.EXPENSE
            amount = income if income > 0 else expense
            CashEntry.objects.get_or_create(
                date=date,
                description=str(description).strip(),
                kind=kind,
                amount=amount,
            )

        self.stdout.write(self.style.SUCCESS("Importação concluída."))
